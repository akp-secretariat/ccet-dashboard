#!/usr/bin/env python3
"""Build small browser-ready CCET dashboard datasets from the NICCDIES Excel workbook.

The workbook has different column layouts by fiscal year, so this script normalizes
known FY2017-FY2026 sheets into one common schema and writes aggregated JSON files.

Amounts are retained in the workbook's published unit: thousand Philippine pesos.
"""
from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import load_workbook

SHEET_LAYOUTS: Dict[str, Dict[str, Any]] = {
    "2017 (Actual)": dict(start=5, dept=0, agency=1, pap_code=2, pap_desc=3, cc_code=4, cc_desc=None, adaptation=5, mitigation=6, total=7),
    "2018 (Actual)": dict(start=5, dept=0, agency=2, pap_code=None, pap_desc=3, cc_code=4, cc_desc=None, adaptation=5, mitigation=6, total=7),
    "2019 (GAA)": dict(start=7, dept=1, agency=3, pap_code=4, pap_desc=5, cc_code=None, cc_desc=None, adaptation=6, mitigation=7, total=8),
    "2019 (Actual)": dict(start=8, dept=2, agency=4, pap_code=5, pap_desc=6, cc_code=7, cc_desc=8, adaptation=9, mitigation=10, total=11),
    "2020 (NEP)": dict(start=7, dept=1, agency=3, pap_code=4, pap_desc=5, cc_code=None, cc_desc=None, adaptation=6, mitigation=7, total=8),
    "2020 (Actual)": dict(start=9, dept=2, agency=4, pap_code=5, pap_desc=6, cc_code=7, cc_desc=8, adaptation=9, mitigation=10, total=11),
    "2021 (Actual)": dict(start=10, dept=1, agency=3, pap_code=4, pap_desc=5, cc_code=6, cc_desc=7, adaptation=8, mitigation=9, total=10),
    "2022 (GAA)": dict(start=9, dept=1, agency=3, pap_code=4, pap_desc=5, cc_code=6, cc_desc=7, adaptation=8, mitigation=9, total=10),
    "2022 (Actual)": dict(start=9, dept=2, agency=4, pap_code=5, pap_desc=6, cc_code=7, cc_desc=8, adaptation=9, mitigation=10, total=11),
    "2023 (NEP)": dict(start=10, dept=1, agency=3, pap_code=4, pap_desc=5, cc_code=6, cc_desc=7, adaptation=8, mitigation=9, total=10),
    "2023 (GAA)": dict(start=9, dept=2, agency=4, pap_code=5, pap_desc=6, cc_code=7, cc_desc=8, adaptation=9, mitigation=10, total=11),
    "2023 (Actual)": dict(start=8, dept=2, agency=4, pap_code=5, pap_desc=6, cc_code=7, cc_desc=8, adaptation=9, mitigation=10, total=11),
    "2024 (NEP)": dict(start=9, dept=2, agency=4, pap_code=5, pap_desc=6, cc_code=7, cc_desc=8, adaptation=9, mitigation=10, total=11),
    "2024 (GAA)": dict(start=8, dept=2, agency=4, pap_code=5, pap_desc=6, cc_code=7, cc_desc=8, adaptation=9, mitigation=10, total=11),
    "2024 (Actual)": dict(start=7, dept=1, agency=3, pap_code=4, pap_desc=5, cc_code=6, nccap=7, cc_desc=8, adaptation=12, mitigation=13, total=14),
    "2025 (NEP)": dict(start=8, dept=2, agency=4, pap_code=5, pap_desc=6, cc_code=7, cc_desc=8, adaptation=9, mitigation=10, total=11),
    "2025 (GAA)": dict(start=8, dept=1, agency=3, pap_code=4, pap_desc=5, cc_code=6, cc_desc=7, adaptation=28, mitigation=29, total=30),
    "2026 (NEP)": dict(start=7, dept=1, agency=3, pap_code=4, pap_desc=5, cc_code=6, nccap=7, cc_desc=8, adaptation=16, mitigation=17, total=18),
    "2026 (GAA": dict(start=7, dept=1, agency=3, pap_code=4, pap_desc=5, cc_code=6, nccap=7, cc_desc=8, adaptation=15, mitigation=16, total=17),
}

PROJECT_TYPES: List[Tuple[str, List[str]]] = [
    ("Flood control & drainage", [r"flood", r"drainage", r"dike", r"dyke", r"river control", r"revetment", r"waterway", r"culvert"]),
    ("Coastal protection", [r"seawall", r"sea wall", r"coastal protection", r"storm surge", r"breakwater", r"shore protection", r"coastline"]),
    ("Roads & bridges", [r"\broad\b", r"highway", r"bridge", r"pavement", r"causeway", r"bypass", r"flyover", r"interchange"]),
    ("Water supply & irrigation", [r"irrigation", r"water supply", r"water system", r"dam\b", r"reservoir", r"potable water", r"waterworks"]),
    ("Buildings & public facilities", [r"building", r"school", r"hospital", r"health center", r"evacuation center", r"government center", r"facility", r"warehouse"]),
    ("Ports & transport", [r"port\b", r"airport", r"railway", r"railroad", r"transport terminal", r"seaport"]),
    ("Solid waste & wastewater", [r"solid waste", r"wastewater", r"sewer", r"septage", r"sanitation", r"materials recovery facility", r"landfill"]),
    ("Energy infrastructure", [r"solar", r"wind farm", r"power plant", r"microgrid", r"transmission", r"electrification", r"renewable energy", r"hydropower", r"geothermal"]),
    ("Agriculture & fisheries infrastructure", [r"farm-to-market", r"fish port", r"post-harvest", r"greenhouse", r"cold storage", r"fishpond"]),
    ("Ecosystem / nature-based", [r"mangrove", r"reforestation", r"afforestation", r"watershed", r"forest", r"wetland", r"riverbank vegetation", r"slope protection.*coco", r"bioengineering"]),
]

INFRA_HINTS = re.compile(
    r"construction|rehabilitation|repair|retrof|improvement|upgrading|installation|facility|infrastructure|structure|building|road|bridge|flood|drainage|irrigation|water system|seawall|port|airport|dam|dike|revetment|slope protection|renewable energy|solar|power plant",
    re.I,
)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def num(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return 0.0


def classify_project(pap_desc: str, cc_desc: str) -> str:
    text = f"{pap_desc} {cc_desc}".lower()
    for label, patterns in PROJECT_TYPES:
        if any(re.search(p, text, re.I) for p in patterns):
            return label
    if INFRA_HINTS.search(text):
        return "Other infrastructure"
    return "Non-infrastructure"


def parse_sheet_name(name: str) -> Tuple[Optional[int], str]:
    m = re.search(r"(20\d{2})\s*\((Actual|GAA|NEP)", name, re.I)
    if not m:
        return None, ""
    return int(m.group(1)), m.group(2).upper() if m.group(2).upper() != "ACTUAL" else "Actual"


def get_cell(values: List[Any], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(values):
        return None
    return values[idx]


def load_records(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    records: List[Dict[str, Any]] = []
    missing_layouts = []

    for sheet_name in wb.sheetnames:
        year, stage = parse_sheet_name(sheet_name)
        if not year:
            continue
        layout = SHEET_LAYOUTS.get(sheet_name)
        if not layout:
            missing_layouts.append(sheet_name)
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=layout["start"], values_only=True):
            values = list(row)
            dept = clean_text(get_cell(values, layout.get("dept")))
            agency = clean_text(get_cell(values, layout.get("agency")))
            pap_desc = clean_text(get_cell(values, layout.get("pap_desc")))
            cc_code = clean_text(get_cell(values, layout.get("cc_code")))
            cc_desc = clean_text(get_cell(values, layout.get("cc_desc")))
            pap_code = clean_text(get_cell(values, layout.get("pap_code")))
            nccap = clean_text(get_cell(values, layout.get("nccap")))
            adaptation = num(get_cell(values, layout.get("adaptation")))
            mitigation = num(get_cell(values, layout.get("mitigation")))
            total = num(get_cell(values, layout.get("total")))
            if total == 0 and (adaptation or mitigation):
                total = adaptation + mitigation

            # Skip headers, subtotals and empty lines. A valid record must have an amount
            # or a recognizable department/agency/PAP description.
            if not any([dept, agency, pap_desc, cc_code, adaptation, mitigation, total]):
                continue
            if dept.lower() in {"department", "department name", "uacs dpt dsc"}:
                continue
            # Totals/subtotals in several sheets are display rows, not PAP records.
            if not dept and not agency:
                continue
            if dept.strip().upper() in {"TOTAL", "SUBTOTAL", "SUB-TOTAL", "GRAND TOTAL"}:
                continue
            if pap_desc.strip().upper() in {"TOTAL", "SUBTOTAL", "SUB-TOTAL", "GRAND TOTAL"}:
                continue

            record = {
                "year": year,
                "stage": stage,
                "department": dept,
                "agency": agency,
                "pap_code": pap_code,
                "pap_description": pap_desc,
                "cc_code": cc_code,
                "cc_description": cc_desc,
                "nccap_priority": nccap,
                "adaptation": round(adaptation, 4),
                "mitigation": round(mitigation, 4),
                "total": round(total, 4),
                "project_type": classify_project(pap_desc, cc_desc),
                "province": "",
            }
            records.append(record)

    if missing_layouts:
        print("WARNING: sheets skipped because their layout is not yet configured:", ", ".join(missing_layouts), file=sys.stderr)
    return records


def normalize_place_name(name: str) -> str:
    n = name.lower()
    n = re.sub(r"\b(province of|city of|municipality of|province|city|municipality)\b", " ", n)
    n = n.replace("ñ", "n")
    n = re.sub(r"[^a-z0-9 ]+", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def flatten_coords(coords: Any) -> Iterable[Tuple[float, float]]:
    if not isinstance(coords, list):
        return
    if len(coords) >= 2 and isinstance(coords[0], (int, float)) and isinstance(coords[1], (int, float)):
        yield (float(coords[0]), float(coords[1]))
    else:
        for item in coords:
            yield from flatten_coords(item)


def fetch_places() -> List[Dict[str, Any]]:
    """Fetch Philippines ADM2 boundaries from geoBoundaries and retain only names/centres.

    This avoids shipping a large GeoJSON to the browser. If the service is unavailable,
    the dashboard still builds; only the map will have no inferred points.
    """
    api = "https://www.geoboundaries.org/api/current/gbOpen/PHL/ADM2/"
    try:
        meta = requests.get(api, timeout=30).json()
        gj_url = meta.get("gjDownloadURL")
        if not gj_url:
            return []
        gj = requests.get(gj_url, timeout=90).json()
    except Exception as exc:
        print(f"WARNING: geoBoundaries download failed: {exc}", file=sys.stderr)
        return []

    places: List[Dict[str, Any]] = []
    for feature in gj.get("features", []):
        props = feature.get("properties", {})
        name = props.get("shapeName") or props.get("NAME_2") or props.get("name")
        if not name:
            continue
        pts = list(flatten_coords(feature.get("geometry", {}).get("coordinates", [])))
        if not pts:
            continue
        lons = [p[0] for p in pts]
        lats = [p[1] for p in pts]
        # Bounding-box centre is good enough for a dashboard bubble marker.
        lon = (min(lons) + max(lons)) / 2
        lat = (min(lats) + max(lats)) / 2
        places.append({"name": clean_text(name), "key": normalize_place_name(name), "lat": round(lat, 5), "lon": round(lon, 5)})
    return places


def infer_locations(records: List[Dict[str, Any]], places: List[Dict[str, Any]]) -> None:
    # Longest names first prevents e.g. "Davao" swallowing "Davao del Sur".
    candidates = sorted(places, key=lambda p: len(p["key"]), reverse=True)
    for r in records:
        if r["project_type"] == "Non-infrastructure":
            continue
        text = normalize_place_name(f"{r['pap_description']} {r['agency']}")
        padded = f" {text} "
        for p in candidates:
            key = p["key"]
            if key and f" {key} " in padded:
                r["province"] = p["name"]
                break


def aggregate(records: List[Dict[str, Any]], dims: List[str]) -> List[Dict[str, Any]]:
    acc: Dict[Tuple[Any, ...], List[float]] = defaultdict(lambda: [0.0, 0.0, 0.0, 0.0])
    for r in records:
        key = tuple(r[d] for d in dims)
        a = acc[key]
        a[0] += r["adaptation"]
        a[1] += r["mitigation"]
        a[2] += r["total"]
        a[3] += 1
    out = []
    for key, vals in acc.items():
        item = {d: key[i] for i, d in enumerate(dims)}
        item.update(adaptation=round(vals[0], 4), mitigation=round(vals[1], 4), total=round(vals[2], 4), records=int(vals[3]))
        out.append(item)
    return out


def write_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def download_source(url: str) -> Path:
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    Path(tmp).write_bytes(r.content)
    return Path(tmp)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", help="Local .xlsx path. If omitted, CCET_XLSX_URL is downloaded.")
    ap.add_argument("--out", default="site/data", help="Output data directory")
    args = ap.parse_args()

    temp_path: Optional[Path] = None
    if args.input:
        source = Path(args.input)
    else:
        url = os.environ.get("CCET_XLSX_URL", "").strip()
        if not url:
            print("ERROR: set repository variable CCET_XLSX_URL to the direct .xlsx URL, or pass --input.", file=sys.stderr)
            return 2
        print(f"Downloading CCET workbook from {url}")
        temp_path = download_source(url)
        source = temp_path

    if not source.exists():
        print(f"ERROR: source workbook not found: {source}", file=sys.stderr)
        return 2

    print(f"Reading {source}")
    records = load_records(source)
    print(f"Normalized {len(records):,} records")

    places = fetch_places()
    if places:
        infer_locations(records, places)
    print(f"Location reference points: {len(places):,}")

    out = Path(args.out)
    write_json(out / "year_stage.json", aggregate(records, ["year", "stage"]))
    write_json(out / "departments.json", aggregate(records, ["year", "stage", "department"]))
    write_json(out / "agencies.json", aggregate(records, ["year", "stage", "department", "agency"]))
    write_json(out / "nccap.json", aggregate([r for r in records if r["nccap_priority"]], ["year", "stage", "nccap_priority"]))
    write_json(out / "typologies.json", aggregate([r for r in records if r["cc_code"]], ["year", "stage", "cc_code", "cc_description"]))
    write_json(out / "project_types.json", aggregate(records, ["year", "stage", "project_type"]))
    write_json(out / "map.json", aggregate([r for r in records if r["province"] and r["project_type"] != "Non-infrastructure"], ["year", "stage", "province", "project_type"]))
    write_json(out / "places.json", places)

    # Top PAPs are limited to keep browser payload small while still allowing useful drill-down.
    top = sorted(records, key=lambda r: r["total"], reverse=True)[:3000]
    keep = ["year", "stage", "department", "agency", "pap_code", "pap_description", "cc_code", "cc_description", "nccap_priority", "adaptation", "mitigation", "total", "project_type", "province"]
    write_json(out / "top_projects.json", [{k: r[k] for k in keep} for r in top])

    years = sorted({r["year"] for r in records})
    stages_by_year = {str(y): sorted({r["stage"] for r in records if r["year"] == y}) for y in years}
    metadata = {
        "source_file": source.name if args.input else os.environ.get("CCET_XLSX_URL", ""),
        "source_unit": "thousand Philippine pesos",
        "normalized_records": len(records),
        "years": years,
        "stages_by_year": stages_by_year,
        "location_method": "Province/city names are inferred from PAP/agency text and matched to geoBoundaries ADM2 names. Unmatched projects are not shown on the map.",
        "geo_attribution": "Administrative reference names/centres: geoBoundaries gbOpen (CC BY 4.0).",
    }
    write_json(out / "metadata.json", metadata)

    if temp_path:
        try:
            temp_path.unlink()
        except OSError:
            pass
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
